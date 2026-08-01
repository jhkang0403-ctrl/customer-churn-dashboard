"""data_messy_상담이력.xlsx 정제 스크립트 (정제_규칙_상담이력.md 기준).

정제 규칙 순서:
    0. 재문의여부(I열) 색상은 pandas로 읽을 수 없으므로, openpyxl로 셀 값+색을 함께 읽는다.
    1. 월별 시트를 하나로 합치되, 시트명을 "월" 컬럼으로 남긴다 (2행 헤더 병합 포함).
    2. 병합 셀(소속)을 그룹 내 forward fill로 채운다.
    3. 소계 행을 분리한다 (상담원에 "소계" 포함 여부).
    4. 상담일(3가지 혼재 형식)을 ISO(YYYY-MM-DD)로 통일하고, 월 컬럼과 일치하는지 검증한다.
    5. 상담시간을 분 단위 숫자로 통일한다.
    6. 재문의여부를 불리언으로 만든다 (텍스트 "N"->False, 텍스트 없음+연분홍색->True).
    7. 소계 행 값(처리결과=그룹 건수, 재문의여부=True 건수)과 그룹 집계를 대조 검증한 뒤 소계 행을 제거한다.
"""
import re
from datetime import date, timedelta
from pathlib import Path

import openpyxl
import pandas as pd

BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "data"
REPORT_DIR = BASE_DIR / "report"

RAW_PATH = DATA_DIR / "data_messy_상담이력.xlsx"
CLEAN_PATH = DATA_DIR / "data_clean_상담이력.csv"
REPORT_PATH = REPORT_DIR / "11_상담이력_정제리포트.md"

TEAM_COL = "소속"
AGENT_COL = "상담원"
DATE_COL = "상담일"
CUSTOMER_COL = "고객ID"
CHANNEL_COL = "채널"
CATEGORY_COL = "문의유형"
DURATION_COL = "상담시간"
STATUS_COL = "처리결과"
REINQUIRY_COL = "재문의여부"
MONTH_COL = "월"

SUBTOTAL_KEYWORD = "소계"
REINQUIRY_PINK_FILL = "FFFFC7CE"
EXCEL_EPOCH = date(1899, 12, 30)


def _build_column_names(ws) -> list[str]:
    """2행 헤더(상위/하위)를 하나의 컬럼명으로 합친다.

    하위 행(2행)에 값이 있으면 그 값을, 없으면 상위 행(1행) 값을 컬럼명으로 쓴다.
    "상담 결과"(H1:I1 병합) 아래 "처리결과"/"재문의여부"만 하위 값을 쓰고,
    나머지 컬럼은 상위 값(세로 병합 라벨)을 그대로 쓴다.
    """
    names = []
    for col in range(1, ws.max_column + 1):
        top = ws.cell(row=1, column=col).value
        bottom = ws.cell(row=2, column=col).value
        names.append(bottom if bottom not in (None, "") else top)
    return names


def load_raw_sheets(path: Path) -> pd.DataFrame:
    """모든 시트를 openpyxl로 읽어 월 컬럼을 붙이고 하나의 데이터프레임으로 합친다.

    재문의여부(I열)는 텍스트뿐 아니라 배경 채우기 색도 함께 읽어야 하므로
    pd.read_excel이 아니라 openpyxl로 셀 단위로 읽어, 색 정보를 "_재문의_fill"
    이라는 내부 컬럼에 별도로 담아둔다.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    columns = None
    reinquiry_idx = None
    frames = []

    for ws in wb.worksheets:
        if columns is None:
            columns = _build_column_names(ws)
            reinquiry_idx = columns.index(REINQUIRY_COL) + 1

        rows = []
        for r in range(3, ws.max_row + 1):
            record = {}
            for c, name in enumerate(columns, start=1):
                record[name] = ws.cell(row=r, column=c).value
            record["_재문의_fill"] = ws.cell(row=r, column=reinquiry_idx).fill.fgColor.rgb
            record[MONTH_COL] = ws.title
            rows.append(record)

        frames.append(pd.DataFrame(rows))

    return pd.concat(frames, ignore_index=True)


def fill_merged_team_cells(df: pd.DataFrame, team_col: str = TEAM_COL) -> pd.DataFrame:
    """규칙 2: 병합 셀 때문에 비어 있는 소속 값을 위 값으로 채운다."""
    df = df.copy()
    df[team_col] = df[team_col].ffill()
    return df


def split_subtotal_rows(
    df: pd.DataFrame, agent_col: str = AGENT_COL, keyword: str = SUBTOTAL_KEYWORD
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """규칙 3: 상담원명에 "소계" 키워드가 포함된 행을 개별 상담 기록과 분리한다."""
    is_subtotal = df[agent_col].astype(str).str.contains(keyword, na=False)
    df_subtotal = df[is_subtotal].copy()
    df_detail = df[~is_subtotal].copy()
    return df_detail, df_subtotal


def _parse_consult_date(value) -> tuple[str | None, str]:
    """상담일 원본 값을 ISO(YYYY-MM-DD) 문자열로 변환한다. (변환값, 원본형식이름) 반환."""
    if isinstance(value, int):
        parsed = EXCEL_EPOCH + timedelta(days=value)
        return parsed.strftime("%Y-%m-%d"), "raw_serial"
    text = str(value).strip()
    for fmt, kind in (("%Y-%m-%d", "YYYY-MM-DD"), ("%y.%m.%d", "YY.M.D")):
        try:
            parsed = pd.to_datetime(text, format=fmt).date()
            return parsed.strftime("%Y-%m-%d"), kind
        except ValueError:
            continue
    return None, "UNPARSED"


def convert_consult_date(df: pd.DataFrame, date_col: str = DATE_COL, month_col: str = MONTH_COL) -> pd.DataFrame:
    """규칙 4: 상담일 3종 혼재 형식을 ISO로 통일하고, 시트월(월 컬럼)과 일치하는지 검증한다.

    일치하지 않는 값이 있어도 강제로 고치지 않고 원본을 남긴 채 경고만 남긴다
    (오기인지 실제로 다른 달 데이터인지는 사람이 판단해야 하는 영역).
    """
    df = df.copy()
    parsed = df[date_col].map(_parse_consult_date)
    df[date_col] = parsed.map(lambda t: t[0])
    df["_날짜_원본형식"] = parsed.map(lambda t: t[1])
    df["_날짜_월불일치"] = df[date_col].str.slice(0, 7) != df[month_col]
    return df


def convert_duration_minutes(df: pd.DataFrame, duration_col: str = DURATION_COL) -> pd.DataFrame:
    """규칙 5: "16분" 같은 문자열과 순수 숫자가 섞인 상담시간을 분 단위 정수로 통일한다."""
    df = df.copy()
    df[duration_col] = (
        df[duration_col].astype(str).str.replace(r"[^0-9]", "", regex=True).astype(int)
    )
    return df


def add_is_recontact_column(
    df: pd.DataFrame, reinquiry_col: str = REINQUIRY_COL
) -> pd.DataFrame:
    """규칙 6: 재문의여부를 불리언으로 만든다.

    텍스트 "N" -> False, 텍스트 없음+연분홍색(FFFFC7CE) -> True.
    원본에는 없지만 텍스트 "Y"가 있다면 True로 처리한다(대비용).
    세 가지 어느 쪽에도 해당하지 않는 값이 있으면 명시적으로 표시해 사람이 확인하게 한다.
    """
    df = df.copy()
    text = df[reinquiry_col].fillna("").astype(str).str.strip()
    is_pink = df["_재문의_fill"] == REINQUIRY_PINK_FILL

    is_recontact = pd.Series(pd.NA, index=df.index, dtype="object")
    is_recontact[text == "N"] = False
    is_recontact[text == "Y"] = True
    is_recontact[(text == "") & is_pink] = True

    unresolved = is_recontact.isna()
    if unresolved.any():
        print(f"  [경고] 재문의여부를 판정할 수 없는 행 {unresolved.sum()}건 발견 (원본 텍스트/색상 확인 필요)")

    df[reinquiry_col] = is_recontact.astype("boolean")
    return df


def cross_check_group_totals(
    df_detail: pd.DataFrame,
    df_subtotal: pd.DataFrame,
    month_col: str = MONTH_COL,
    team_col: str = TEAM_COL,
    status_col: str = STATUS_COL,
    reinquiry_col: str = REINQUIRY_COL,
) -> list[dict]:
    """제거할 소계 행 값과 팀별 실제 집계가 일치하는지 검증한다.

    처리결과(H) 소계는 "그룹 총 상담 건수", 재문의여부(I) 소계는
    "그룹 내 재문의(True) 건수"를 의미한다 (Confirm 단계에서 확인한 정의).
    """
    grouped = (
        df_detail.groupby([month_col, team_col], sort=False)
        .agg(건수=(status_col, "count"), 재문의건수=(reinquiry_col, "sum"))
        .reset_index()
    )

    results = []
    for _, row in df_subtotal.iterrows():
        month, team = row[month_col], row[team_col]
        subtotal_count = int(re.sub(r"[^0-9]", "", str(row[status_col])))
        subtotal_reinquiry = int(re.sub(r"[^0-9]", "", str(row[reinquiry_col])))

        match = grouped[(grouped[month_col] == month) & (grouped[team_col] == team)]
        if match.empty:
            results.append({"월": month, "소속": team, "status": "NO MATCH FOUND"})
            continue
        calc = match.iloc[0]
        results.append(
            {
                "월": month,
                "소속": team,
                "건수_원본": subtotal_count,
                "건수_계산": int(calc["건수"]),
                "건수_일치": subtotal_count == int(calc["건수"]),
                "재문의_원본": subtotal_reinquiry,
                "재문의_계산": int(calc["재문의건수"]),
                "재문의_일치": subtotal_reinquiry == int(calc["재문의건수"]),
            }
        )
    return results


def main() -> None:
    report = []

    def log(line: str = "") -> None:
        report.append(line)
        print(line)

    df_raw = load_raw_sheets(RAW_PATH)
    rows_before = len(df_raw)
    log(f"원본 행 수(소계 포함): {rows_before}")
    log("=" * 60)

    df = fill_merged_team_cells(df_raw)
    df_detail, df_subtotal = split_subtotal_rows(df)
    log(f"[1] 소계 행 분리: {len(df_subtotal)}건 분리 -> 개별 상담 기록 {len(df_detail)}건")

    df_detail = convert_consult_date(df_detail)
    unparsed = (df_detail["_날짜_원본형식"] == "UNPARSED").sum()
    mismatched = df_detail["_날짜_월불일치"].sum()
    log(f"[2] 상담일 ISO(YYYY-MM-DD) 통일: 형식별 건수 {df_detail['_날짜_원본형식'].value_counts().to_dict()}")
    log(f"     - 파싱 실패: {unparsed}건 / 시트월과 불일치: {mismatched}건")

    df_detail = convert_duration_minutes(df_detail)
    log(f"[3] 상담시간을 분 단위 정수로 통일 완료 (범위 {df_detail[DURATION_COL].min()}~{df_detail[DURATION_COL].max()}분)")

    df_detail = add_is_recontact_column(df_detail)
    recontact_count = int(df_detail[REINQUIRY_COL].sum())
    log(f"[4] 재문의여부 불리언 변환: True {recontact_count}건 / False {int((~df_detail[REINQUIRY_COL]).sum())}건")

    log("=" * 60)
    log("[팀별 합계 교차검증]")
    log("=" * 60)
    check_results = cross_check_group_totals(df_detail, df_subtotal)
    for r in check_results:
        if r.get("status") == "NO MATCH FOUND":
            log(f"  {r['월']} {r['소속']}: 대응하는 소계 행을 찾을 수 없음")
            continue
        c_flag = "OK" if r["건수_일치"] else "MISMATCH"
        i_flag = "OK" if r["재문의_일치"] else "MISMATCH"
        log(
            f"  {r['월']} {r['소속']}: 건수 원본={r['건수_원본']} vs 계산={r['건수_계산']} [{c_flag}] | "
            f"재문의 원본={r['재문의_원본']} vs 계산={r['재문의_계산']} [{i_flag}]"
        )

    df_detail = df_detail.drop(columns=["_재문의_fill", "_날짜_원본형식", "_날짜_월불일치"])

    output_columns = [
        MONTH_COL, TEAM_COL, AGENT_COL, DATE_COL, CUSTOMER_COL,
        CHANNEL_COL, CATEGORY_COL, DURATION_COL, STATUS_COL, REINQUIRY_COL,
    ]
    df_detail = df_detail[output_columns]

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    df_detail.to_csv(CLEAN_PATH, index=False, encoding="utf-8-sig")

    log("=" * 60)
    log("[정제 결과 요약]")
    log("=" * 60)
    log(f"  정제 전 행 수(소계 포함): {rows_before}")
    log(f"  정제 후 행 수: {len(df_detail)}")
    log(f"  저장 경로: {CLEAN_PATH}")

    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    with open(REPORT_PATH, "w", encoding="utf-8") as f:
        f.write("# data_messy_상담이력.xlsx 정제 리포트\n\n```\n")
        f.write("\n".join(report))
        f.write("\n```\n")
    log(f"리포트 저장: {REPORT_PATH}")


if __name__ == "__main__":
    main()
